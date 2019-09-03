from openpyxl import load_workbook
import pandas as pd


def add_to_excel(ws, mark, offset=False, start=0, end=5):
    df = pd.read_csv('/Users/xiahualou/PycharmProjects/logExtract/data/csv/rabbitmq/rabbitmq20190903.csv')
    df.loc[len(df)] = df.columns
    df.columns = range(10)
    consts = list('ABCDEFGHIJKLMNOPQRSTUVWXYZ')
    rows_length = ws.max_row
    for column in range(start, end):
        index = consts[consts.index(mark) + column]
        i = 0
        for row in range(rows_length - 3, rows_length - 1):
            if not offset:
                ws["{}{}".format(index, row)].value = float(df[column][i])
            else:
                ws["{}{}".format(index, row)].value = float(df[column + 5][i])
            i += 1
