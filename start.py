# -*- encoding: utf-8 -*-
import time
import pandas
from openpyxl import load_workbook
from libs import config
from threadsPool.t_pool import ThreadPool
import os
import workers


class Ignition:
    def __init__(self):
        self.pd = pandas
        self.success = []
        self.base_path = os.path.dirname(__file__)
        self.template = os.path.join(self.base_path, 'report.xlsx')
        self.consts = list('ABCDEFGHIJKLMNOPQRSTUVWXYZ')
        if os.path.exists(self.template):
            self.wb = load_workbook(filename=self.template)
        else:
            raise Exception("no report.xlsx")
        if len(config.workers) == 0:
            workers = os.listdir(os.path.join(os.path.dirname(__file__), 'workers'))
            try:
                on_remove = ["__init__.py", 'Basewoker.py', '__pycache__']
                [workers.remove(_) for _ in on_remove]
            except Exception:
                pass
            self.workers = [worker.split('.')[0] for worker in workers]
        else:
            self.workers = config.workers
        self.pool = ThreadPool(config.config.getint('threads', 'maximum'))

    def add_to_excel(self, mic):
        path = os.path.join(self.base_path, "data/csv/{}/{}.csv".format(mic, mic))
        df = self.pd.read_csv(path)
        if len(df.columns) != 10:
            print("{} size less than five,can't write to report.xlxs".format(mic))
            return
        df2 = self.pd.DataFrame([df.columns])
        df.columns = range(10)
        df2 = df2.append(df, ignore_index=True)
        platform = config.config.get("platform", "platform")
        ws = self.wb[mic]
        if platform == "ubuntu":
            self.excel_mapping(ws, df2, "T")
            self.excel_mapping(ws, df2, "L")
        else:
            self.excel_mapping(ws, df2, "D")

    @staticmethod
    def find_end_in_sheet(sheet):
        for column in sheet.iter_cols():
            for cell2 in column:
                if cell2.value is not None:
                    info2 = cell2.value.find('Latency Score (Small is better)')
                    if info2 == 0:
                        return cell2.row

    def excel_mapping(self, ws, df, mark, start=0, end=5):
        rows_length = self.find_end_in_sheet(ws)
        for column in range(start, end):
            index = self.consts[self.consts.index(mark) + column]
            i = 0
            for row in range(4, rows_length - 1):
                if mark == "T":
                    try:
                        ws["{}{}".format(index, row)].value = float(df[column][i])
                    except Exception:
                        ws["{}{}".format(index, row)].value = df[column][i]
                else:
                    try:
                        ws["{}{}".format(index, row)].value = float(df[column + 5][i])
                    except Exception:
                        ws["{}{}".format(index, row)].value = df[column + 5][i]
                i += 1

    def write(self):
        self.wb.save("report.xlsx")

    def save(self, status_map):
        for key in status_map.keys():
            if not status_map[key]:
                continue
            try:
                self.add_to_excel(key.lower())
                print("write {} to report successfully".format(key))
            except Exception as Ex:
                print("failed to write {} to report,please check data/csv/{}.csv and make sure it's right".format(key,
                                                                                                                  key))

    @staticmethod
    def wait(t=1):
        time.sleep(t)

    def run(self):
        for work in self.workers:
            if not hasattr(workers, work):
                continue
            workers.task_status[work] = True
            try:
                runner = getattr(getattr(workers, work), work)()
                setattr(runner, '{}_container'.format(work.lower()), [])
            except Exception as Ex:
                workers.task_status[work] = False
                print("task {}.log done Status: extract failed,Error: {}\n".format(work, Ex))
            else:
                self.pool.run(func=runner.run, args=())
        self.pool.close()
        self.pool.join()
        self.save(workers.task_status)
        self.write()


if __name__ == '__main__':
    Ignition().run()
